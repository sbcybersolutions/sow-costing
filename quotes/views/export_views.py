from django.http import HttpResponse
from django.template.loader import render_to_string
from weasyprint import HTML
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from quotes.models import Quote, Talent

def export_excel(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index') #type: ignore

    quote = Quote.objects.get(pk=quote_id)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoW Detailed Costs" #type: ignore

    ws.append(["Asset Type", "Description", "Internal Cost", "Retail Cost"]) #type: ignore

    for course in quote.courses.all(): # type: ignore
        ws.append(["Course", course.description, f"${course.get_total_internal_cost():.2f}", f"${course.get_total_retail_cost():.2f}"]) # type: ignore

    for lv in quote.live_videos.all(): # type: ignore
        ws.append(["Live Video", lv.description, f"${lv.get_total_internal_cost():.2f}", f"${lv.get_total_retail_cost():.2f}"]) # type: ignore
        for talent in lv.talents.all():
            ws.append(["  └ Talent", f"{talent.name} ({talent.role_type})", f"${talent.get_internal_cost():.2f}", f"${talent.get_retail_cost():.2f}"]) # type: ignore

    for av in quote.animated_videos.all(): # type: ignore
        ws.append(["Animated Video", av.description, f"${av.get_total_internal_cost():.2f}", f"${av.get_total_retail_cost():.2f}"]) # type: ignore

    for studio in quote.studios.all(): # type: ignore
        ws.append(["Studio", studio.studio_name, f"${studio.get_total_internal_cost():.2f}", f"${studio.get_total_retail_cost():.2f}"]) # type: ignore

    for tech in quote.technical_staff.all(): # type: ignore
        ws.append(["Technical Staff", # type: ignore
                   f"{tech.filming_days} filming day(s), {tech.editing_days} editing day(s)",
                   f"${tech.get_total_internal_cost():.2f}", f"${tech.get_total_retail_cost():.2f}"])

    for col in ws.columns: # type: ignore
        max_length = 0
        col_letter = get_column_letter(col[0].column) # type: ignore
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2 # type: ignore

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = 'attachment; filename="SoW_Detailed.xlsx"'
    return response


def export_pdf(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index') #type: ignore

    quote = Quote.objects.get(pk=quote_id)

    total_internal = sum([
        sum(c.get_total_internal_cost() for c in quote.courses.all()), # type: ignore
        sum(lv.get_total_internal_cost() for lv in quote.live_videos.all()), # type: ignore
        sum(av.get_total_internal_cost() for av in quote.animated_videos.all()), # type: ignore
        sum(s.get_total_internal_cost() for s in quote.studios.all()), # type: ignore
        sum(t.get_total_internal_cost() for t in quote.technical_staff.all()), # type: ignore
    ])
    total_retail = total_internal * 2

    context = {
        "client_name": quote.client_name,
        "project_name": quote.project_name,
        "date": quote.date,
        "status": quote.status,
        "courses": quote.courses.all(), # type: ignore
        "live_videos": quote.live_videos.all(), # type: ignore
        "talents": Talent.objects.filter(live_video__quote=quote), # type: ignore
        "animated_videos": quote.animated_videos.all(), # type: ignore
        "studios": quote.studios.all(), # type: ignore
        "technical_staff": quote.technical_staff.all(), # type: ignore
        "total_retail": total_retail,
    }

    html_string = render_to_string('quotes/quote_pdf.html', context)
    pdf_file = HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="SoW_Client.pdf"'
    return response