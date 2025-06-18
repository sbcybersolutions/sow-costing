from django.shortcuts import render, redirect
from .forms import (
    CourseForm, LiveVideoForm, TalentForm,
    AnimatedVideoForm, StudioForm, TechnicalStaffForm
)
from .models import Course, LiveVideo, Talent, AnimatedVideo, Studio, TechnicalStaff

import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from openpyxl.styles import numbers
from quotes.models import (
    Course, CourseResource, LiveVideo, Talent, AnimatedVideo,
    Studio, TechnicalStaff, TechnicalRate, VideoTypeRate, FixedCost, StudioRate, Quote
)


from quotes.models import Quote

def index(request):
    if request.method == 'POST':
        client_name = request.POST['client_name']
        project_name = request.POST['project_name']
        date = request.POST['date']

        # ✅ CREATE THE QUOTE FIRST:
        quote = Quote.objects.create(
            client_name=client_name,
            project_name=project_name,
            date=date,
            created_by=request.user if request.user.is_authenticated else None
        )

        # ✅ Now you can store its ID:
        request.session['quote_id'] = quote.id #type: ignore

        return redirect('builder')

    return render(request, 'quotes/index.html')


def builder(request):
    # ✅ 1) Get active Quote ID safely
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index')

    quote = Quote.objects.get(pk=quote_id)

    # ✅ 2) Blank forms for GET
    course_form = CourseForm()
    live_video_form = LiveVideoForm()
    talent_form = TalentForm(quote=quote)  # Pass quote to TalentForm
    animated_video_form = AnimatedVideoForm()
    studio_form = StudioForm()
    technical_form = TechnicalStaffForm()

    # ✅ 3) Handle POST for each form
    if request.method == 'POST':
        # One helper to reduce repetition:
        def handle_form(form_class, button_name):
            if button_name in request.POST:
                form = form_class(request.POST)
                if form.is_valid():
                    obj = form.save(commit=False)
                    obj.quote = quote
                    obj.save()
                    return True
            return False

        if handle_form(CourseForm, 'add_course'):
            return redirect('builder')
        if handle_form(LiveVideoForm, 'add_live_video'):
            return redirect('builder')
        if handle_form(TalentForm, 'add_talent'):
            return redirect('builder')
        if handle_form(AnimatedVideoForm, 'add_animated_video'):
            return redirect('builder')
        if handle_form(StudioForm, 'add_studio'):
            return redirect('builder')
        if handle_form(TechnicalStaffForm, 'add_technical'):
            return redirect('builder')

    # ✅ 4) Only assets linked to this Quote
    context = {
        'quote': quote,
        'courses': quote.courses.all(), # type: ignore
        'live_videos': quote.live_videos.all(), # type: ignore
        'talents': quote.talents.all(), # type: ignore
        'animated_videos': quote.animated_videos.all(), # type: ignore
        'studios': quote.studios.all(), # type: ignore
        'technical_staff': quote.technical_staff.all(), # type: ignore
        'course_form': course_form,
        'live_video_form': live_video_form,
        'talent_form': talent_form,
        'animated_video_form': animated_video_form,
        'studio_form': studio_form,
        'technical_form': technical_form,
    }
    return render(request, 'quotes/builder.html', context)

def new_quote(request):
    request.session.pop('quote_id', None)
    return redirect('index')

from django.urls import reverse

from django.shortcuts import get_object_or_404

# ✅ Delete a single asset
def delete_item(request, model_name, pk):
    MODEL_MAP = {
        'course': Course,
        'livevideo': LiveVideo,
        'talent': Talent,
        'animatedvideo': AnimatedVideo,
        'studio': Studio,
        'technical': TechnicalStaff,
    }

    Model = MODEL_MAP.get(model_name)
    if Model is None:
        return redirect('builder')

    obj = get_object_or_404(Model, pk=pk)
    obj.delete()
    return redirect('builder')


# ✅ Clear ALL quote items
def clear_all(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index')

    quote = Quote.objects.get(pk=quote_id)
    quote.courses.all().delete() # type: ignore
    quote.live_videos.all().delete() # type: ignore
    quote.animated_videos.all().delete() # type: ignore
    quote.studios.all().delete() # type: ignore
    quote.technical_staff.all().delete() # type: ignore
    quote.talents.all().delete() # type: ignore

    return redirect('builder')


# Dynamic totals
from django.http import JsonResponse

from django.http import JsonResponse
from quotes.models import Quote

def get_totals(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return JsonResponse({'total_internal': 0, 'total_retail': 0})

    quote = Quote.objects.get(pk=quote_id)

    total_internal = sum([
        sum(c.get_total_internal_cost() for c in quote.courses.all()), # type: ignore
        sum(lv.get_total_internal_cost() for lv in quote.live_videos.all()), # type: ignore
        sum(av.get_total_internal_cost() for av in quote.animated_videos.all()), # type: ignore
        sum(s.get_total_internal_cost() for s in quote.studios.all()), # type: ignore
        sum(t.get_total_internal_cost() for t in quote.technical_staff.all()), # type: ignore
    ])

    total_retail = total_internal * 2  # or your custom rule

    return JsonResponse({
        'total_internal': total_internal,
        'total_retail': total_retail
    })


# Export to Excel

def export_excel(request):
    # ✅ 1) Pull active Quote ID from session
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index')

    quote = Quote.objects.get(pk=quote_id)

    # ✅ 2) Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SoW Detailed Costs" # type: ignore

    # ✅ 3) Add header row
    ws.append(["Asset Type", "Description", "Internal Cost", "Retail Cost"]) # type: ignore

    # ✅ 4) Add Courses for this Quote only
    for course in quote.courses.all(): # type: ignore
        ws.append([ # type: ignore
            "Course",
            course.description,
            f"${course.get_total_internal_cost():.2f}",
            f"${course.get_total_retail_cost():.2f}"
        ])

    # ✅ 5) Add Live Videos for this Quote only
    for lv in quote.live_videos.all(): # type: ignore
        ws.append([ #type: ignore
            "Live Video",
            lv.description,
            f"${lv.get_total_internal_cost():.2f}",
            f"${lv.get_total_retail_cost():.2f}"
        ])
        # Include attached Talents for this Live Video
        for talent in lv.talents.all():
            ws.append([ #type: ignore
                "  └ Talent",
                f"{talent.name} ({talent.role_type})",
                f"${talent.get_internal_cost():.2f}",
                f"${talent.get_retail_cost():.2f}"
            ])

    # ✅ 6) Add Animated Videos
    for av in quote.animated_videos.all(): # type: ignore
        ws.append([ # type: ignore
            "Animated Video",
            av.description,
            f"${av.get_total_internal_cost():.2f}",
            f"${av.get_total_retail_cost():.2f}"
        ])

    # ✅ 7) Add Studios
    for studio in quote.studios.all(): # type: ignore
        ws.append([ # type: ignore
            "Studio",
            studio.studio_name,
            f"${studio.get_total_internal_cost():.2f}",
            f"${studio.get_total_retail_cost():.2f}"
        ])

    # ✅ 8) Add Technical Staff
    for tech in quote.technical_staff.all():    # type: ignore
        ws.append([ # type: ignore
            "Technical Staff",
            f"{tech.filming_days} filming day(s), {tech.editing_days} editing day(s)",
            f"${tech.get_total_internal_cost():.2f}",
            f"${tech.get_total_retail_cost():.2f}"
        ])

    # ✅ 9) Auto-fit column widths
    for col in ws.columns: # type: ignore
        max_length = 0
        col_letter = get_column_letter(col[0].column) # type: ignore
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2 # type: ignore

    # ✅ 10) Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="SoW_Detailed.xlsx"'
    return response

# Export to PDF
from django.template.loader import render_to_string
from xhtml2pdf import pisa

def export_pdf(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index')

    quote = Quote.objects.get(pk=quote_id)

    total_internal = sum([
        sum(c.get_total_internal_cost() for c in quote.courses.all()), # type: ignore
        sum(lv.get_total_internal_cost() for lv in quote.live_videos.all()), # type: ignore
        sum(av.get_total_internal_cost() for av in quote.animated_videos.all()), # type: ignore
        sum(s.get_total_internal_cost() for s in quote.studios.all()),  # type: ignore
        sum(t.get_total_internal_cost() for t in quote.technical_staff.all()), # type: ignore
    ])
    total_retail = total_internal * 2

    context = {
        "client_name": quote.client_name,
        "project_name": quote.project_name,
        "date": quote.date,
        "courses": quote.courses.all(), # type: ignore
        "live_videos": quote.live_videos.all(), # type: ignore
        "animated_videos": quote.animated_videos.all(), # type: ignore
        "studios": quote.studios.all(), # type: ignore
        "technical_staff": quote.technical_staff.all(), # type: ignore
        "total_retail": total_retail,
    }

    html = render_to_string('quotes/quote_pdf.html', context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="SoW_Client.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err: #type: ignore
        return HttpResponse('Error generating PDF', status=500)
    return response

from quotes.models import Quote

def quote_review(request):
    quote_id = request.session.get('quote_id')
    if not quote_id:
        return redirect('index')

    quote = Quote.objects.get(pk=quote_id)

    total_internal = sum([
        sum(c.get_total_internal_cost() for c in quote.courses.all()), # type: ignore
        sum(lv.get_total_internal_cost() for lv in quote.live_videos.all()), # type: ignore
        sum(av.get_total_internal_cost() for av in quote.animated_videos.all()), # type: ignore
        sum(s.get_total_internal_cost() for s in quote.studios.all()), # type: ignore
        sum(t.get_total_internal_cost() for t in quote.technical_staff.all()), # type: ignore
    ])

    total_retail = total_internal * 2

    return render(request, 'quotes/quote_review.html', {
        "quote": quote,
        "courses": quote.courses.all(), # type: ignore
        "live_videos": quote.live_videos.all(), # type: ignore
        "animated_videos": quote.animated_videos.all(), # type: ignore
        "studios": quote.studios.all(), # type: ignore
        "technical_staff": quote.technical_staff.all(), # type: ignore
        "total_internal": total_internal,
        "total_retail": total_retail,
    })

